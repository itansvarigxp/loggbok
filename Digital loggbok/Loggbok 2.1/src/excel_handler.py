
from datetime import datetime, date, timedelta
from string import ascii_uppercase
import openpyxl
import member
import paths
from os import makedirs, path
from collections import OrderedDict
import statistics_handler as StatLogger

# Antal dagar tillbaks i tiden som sparas i loggbok online.
days_saved_online = 21
latest_save = datetime.now()
data_logging = True
ascii_sheet = list(ascii_uppercase) + list(map(lambda y : 'A' + y, list(ascii_uppercase))) \
                                    + list(map(lambda y : 'B' + y, list(ascii_uppercase)))

statistics_categories = OrderedDict([('Datum', StatLogger.yesterday),
                         ('Unika besökare idag', StatLogger.uniqueVisitorsToday),
                         ('Unika besökare denna månad', StatLogger.uniqueVisitorsMonth),
                         ('Totalt antal timmar i verkstaden idag', StatLogger.totalTimeTodayStrf),
                         ('Styret antal timmar i verkstaden idag', StatLogger.totalTimeTodayStyretStrf),
                         ('Totalt antal timmar i verkstaden denna månad', StatLogger.totalTimeMonthStrf),
                         ('Styret antal timmar i verkstaden denna månad', StatLogger.totalTimeMonthStyretStrf),
                         ('Antal incheckningar totalt idag', StatLogger.checkinsToday),
                         ('Genomsnittlig daglig verksamhet', StatLogger.dailyMeanStrf),
                         ('Standardavvikelse', StatLogger.dailyStdDev),
                         ('Genomsnittlig tid i verkstaden', StatLogger.monthlyMeanStrf),
                         ('Standardavvikelse', StatLogger.monthlyStdDev),
                         ('Styret genomsnittlig tid i verkstaden', StatLogger.monthlyMeanStyretStrf),
                         ('Styret Standardavvikelse', StatLogger.monthlyStdDevStyret),
                         ('Glömda utcheckningar', StatLogger.forgottenCheckOuts)])


def createNewLoggbook():
    global loggbok
    global loggSheet
    loggbok = openpyxl.Workbook()
    loggSheet = loggbok.active
    loggSheet['A1'] = 'Datum'
    loggSheet['B1'] = 'Namn'
    loggSheet['C1'] = 'Incheckning'
    loggSheet['D1'] = 'Utcheckning'
    loggSheet['E1'] = 'Anmärkning'

# Om det finns en loggbok online så laddas den in, annar skapas en ny.
try:
    loggbok = openpyxl.load_workbook(paths.xlsx_logg_online)
    loggSheet = loggbok.active
except:
    createNewLoggbook()

def exctractYesterdayFromLog():
    yesterday = date.today() - timedelta(1)
    list_of_checkins = []
    for row in range(2, loggSheet.max_row):
        date = datetime.strptime(loggSheet['A'+str(row)].value,"%Y-%m-%d").date()
        if date == yesterday:
            name =loggSheet['B' + row]
            checkin = loggSheet['C' + row]
            checkout = loggSheet['D' + row]
            list_of_checkins.append({'name': name, 'checkin': checkin, 'checkout':checkout})
    return list_of_checkins

# Initierar medlemsregistret. Om den misslyckas så skrivs det ut i konsolen
def initMemberRegister():
    try:
        medreg = openpyxl.load_workbook(paths.xlsx_member_register)
        medSheet = medreg[medreg.sheetnames[0]]
        for row in range(2, medSheet.max_row + 1):
            keyCard = medSheet['A' + str(row)].value
            name =  medSheet['B' + str(row)].value
            board_member = medSheet['C' + str(row)].value == 'Styret'
            latest_activity = medSheet['D' + str(row)].value
            member.Member(keyCard, name, board_member, latest_activity)
    except:
        print("Could not find member register.\n" +
              "Please place the member register in the following folder\n" + 
              paths.xlsx_member_register)


# Rensar loggbok online från inloggningar som är äldre än days_saved_online
def cleanEarliestLoggbook():
    old_loggbook = loggbok
    old_loggSheet = old_loggbook.active
    createNewLoggbook()
    global loggSheet
    index = 2
    for row in range(2, old_loggSheet.max_row + 1):
        time = old_loggSheet['A'+str(row)].value
        if time == None or ((datetime.today() - datetime.strptime(time,"%Y-%m-%d")).days > days_saved_online):
            pass
        else:
            i = str(index)
            loggSheet['A' + i] = old_loggSheet['A'+str(row)].value
            loggSheet['B' + i] = old_loggSheet['B'+str(row)].value
            loggSheet['C' + i] = old_loggSheet['C'+str(row)].value
            loggSheet['D' + i] = old_loggSheet['D'+str(row)].value
            loggSheet['E' + i] = old_loggSheet['E'+str(row)].value
            index += 1
    save()

# Sparar ned loggboken på fil
def save():
        loggbok.save(paths.xlsx_logg_online)
        global latest_save
        latest_save = datetime.now()
        # ERROR ERROR ERROR FIX THIS
        #current_month = datetime.now().strftime('%Y%B')
        #loggbok.save('%s%s.xlsx' %(paths.xlsx_datalogger, current_month))
        

# Parser för att kodläsaren i verkstaden läser 24 bitar medan den på kontoret
# läser 32 bitar. Så länge kodläsaren på kontoret läser mer än 24 bitar så
# parsas denna data korrekt
def cardreaderParser(cardkey_str):
    cardreader_bits = 24
    cardkey_binary = bin(int(cardkey_str))[2:]
    cardkey_binary_appended = '0'*cardreader_bits + cardkey_binary
    return ('0,%i' %int(cardkey_binary_appended[-cardreader_bits:], 2))

# Importerar nya medlemmar från separat register. Om filen inte finns
# så skapas den. När medlemmarna skrivits in i medlemsdatabasen så
# töms filen med nya medlemmar.
def importNewMembers():
    try:
        medreg = openpyxl.load_workbook(paths.xlsx_new_members)
        medSheet = medreg[medreg.sheetnames[0]]
        for row in range(2, medSheet.max_row + 1):
            key_card = cardreaderParser(medSheet['A' + str(row)].value)
            name =  medSheet['B' + str(row)].value
            board_member = medSheet['C' + str(row)].value == 'Styret'
            latest_activity = datetime.now().strftime("%Y-%m-%d")
            member.Member(key_card, name, board_member, latest_activity)
        medreg.close()
    except:
        print('Unable to import new members')
    medreg = openpyxl.Workbook()
    medSheet = medreg.active
    medSheet.title = 'Nya medlemmar'
    medSheet['A1'] = 'Nyckelnr'
    medSheet['B1'] = 'Namn'
    medSheet['C1'] = 'Anmärkning'
    medreg.save(paths.xlsx_new_members)
    medreg.close()
    saveMemberlistToFile()

# Spara den aktiva medlemslistan till fil
def saveMemberlistToFile():
    medreg = openpyxl.Workbook()
    medSheet = medreg.active
    medSheet.title = 'Medlemsregister'
    medSheet['A1'] = '0,Nyckelnr'
    medSheet['B1'] = 'Namn'
    medSheet['C1'] = 'Anmärkning'
    medSheet['D1'] = 'Senast aktivitet'
    row = 2
    for key in member.Member.member_register:
        members = member.Member.member_register[key]
        medSheet['A' + str(row)] = members.getKeyCardNumber()
        medSheet['B' + str(row)] = members.getName()
        if members.getBoardmember():
            medSheet['C' + str(row)] = 'Styret'
        else:
            medSheet['C' + str(row)] = None
        medSheet['D' + str(row)] = members.getLatestActivity()
        row += 1
    medreg.save(paths.xlsx_member_register)
    medreg.close()

# Sparar alla incheckade som glömde checka ut
def saveAllCheckedinToLog():
    global loggSheet
    for key in member.Member.checked_in_members:
        row = str(loggSheet.max_row + 1)
        members = member.Member.checked_in_members[key]
        loggSheet['A' + row] = members.getCheckinDate()
        loggSheet['B' + row] = members.getName()
        loggSheet['C' + row] = members.getCheckedInTime()
        loggSheet['E' + row] = 'Did not checkout'
    for key in member.Member.checked_in_styret:
        row = str(loggSheet.max_row + 1)
        members = member.Member.checked_in_styret[key]
        loggSheet['A' + row] = members.getCheckinDate()
        loggSheet['B' + row] = members.getName()
        loggSheet['C' + row] = members.getCheckedInTime()
        loggSheet['E' + row] = 'Did not checkout'
    save()

# Sparar en medlem i den aktiva loggboken (obs. sparas
# inte till excelfilen automatiskt. Detta görs i funktionen
# save)
def saveToLog(member):
    time_now_str = datetime.now().strftime("%H:%M:%S")
    row = str(loggSheet.max_row + 1)
    loggSheet['A' + row] = member.getCheckinDate()
    loggSheet['B' + row] = member.getName()
    loggSheet['C' + row] = member.getCheckedInTime()
    loggSheet['D' + row] = time_now_str
    # Om medlemmen checkar ut efter 00:00 så noteras det
    if time_now_str < member.getCheckedInTime():
        loggSheet['E' + row] = "Late checkout"

def saveStatistics():
    today = date.today()
    current_date = today.strftime('%m/%d')
    current_month = today.strftime('%b')
    yesterday = date.today() - timedelta(1)
    stat_file_path = ('%s%s.xlsx' %(paths.xlsx_statistics, yesterday.strftime('%Y%B')))
    try:
        statistics = openpyxl.load_workbook(stat_file_path)
        stat_sheet = statistics.active
    except:
        if not path.exists(paths.xlsx_statistics):
            makedirs(paths.xlsx_statistics)
        statistics = openpyxl.Workbook()
        stat_sheet = statistics.active
        stat_sheet.append(['%s' % ordered_key for ordered_key in statistics_categories])
    
    stat_sheet.append(['%s' % statistics_categories[ordered_key]() for ordered_key in statistics_categories])
    if StatLogger.unique_visitors_this_month['CURRENTMONTH'] != current_month:
        StatLogger.resetMonth()
    if StatLogger.unique_visitors_today['CURRENTDAY'] != current_date:
        StatLogger.resetToday()
    statistics.save(stat_file_path)
    statistics.close()


