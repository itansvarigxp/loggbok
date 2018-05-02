import openpyxl, datetime, threading, time
from threading import Thread, Timer
from tkinter import *
from pyautogui import press, typewrite
from shutil import copyfile
from os import remove
#from PIL import Image, ImageTK

#sudo apt-get install libtiff4-dev libjpeg8-dev zlib1g-dev libfreetype6-dev liblcms2-dev libwebp-dev tcl8.5-dev tk8.5-dev python-tk

#sudo apt-get install python3-xlib python3-pip python3.tk
#sudo pip3 install pyautogui threading openpyxl datetime


# Variabler som används
card_number = ''
exit_flag = False


checking = ''
logg = 0
unknown = 'to XP'
namn = unknown
oldDate = []
status = False
incheckade = []
window = 0
reset = False
Styret = False
Styret_antal = 0

root = Tk()
root.title("Incheckade XP")
alpha = False
regkort = 0

# Bakgrundsloop som uppdaterar loggboken vid behov
# samt tömmer loggboken vid ett visst klockslag
def bg_main():
    while True:
        today = datetime.datetime.now()
        time_now = today.time().strftime("%H:%M:%S")  # kolla klockan  
        if time_now >= ('04:00:00') and time_now <= ('05:00:00'):  # Mellan 4 & 5
            clear():
        save()
        time.sleep(3599)

t_bgmain = Thread(target=bg_main)
t_bgmain.start()

def clear():
    incheckade = []
    Styret_antal = 0
    reset = True

# Time-out funktion för byte av kort

def abort():
    typewrite('Aborted!')
    press('enter')

# Spara listan externt loggbok.save('info/Loggbok_extern.xlsx')
def save():
        loggbok.save('info/Loggbok_extern.xlsx')

def init_medlemsregister():
    medreg = openpyxl.load_workbook('Medlemsregister.xlsx')
    medSheet = medreg.get_sheet_by_name('Sheet1')
    loggbok = openpyxl.load_workbook('Loggbok.xlsx')
    loggSheet = loggbok.active


def exit():
    exit_flag = True

commands = {
    'exit' : exit,
    'clear' : clear,
    'update' : update
}

# Starta huvudloopen
def main():
    init_medlemsregister()

    while !exit_flag:
        # Wait for input from NFC reader
        card_number = input("Please scan your card now!  ")
        
        if card_number in commands:
            pass
        else:
            card_number = "0," + card_number

        # Number was read
        today = datetime.datetime.now()
        date = today.date()
        year = date.year
        month = today.month
        day = today.day
        time_now = today.time().strftime("%H:%M:%S")





while checking != '0,exit':

    # Vanta pa input fran lasare     
    checking1 = input("Please scan your card now!  ")
    checking = "0,"+checking1

    # Registrera datum och tid
    today = datetime.datetime.now()
    date = today.date()
    year = date.year
    month = today.month
    day = today.day
    time_now = today.time().strftime("%H:%M:%S")
    
    # Om listan ska tömmas
    if checking == '0,Rensa listan':
        incheckade = []
        Styret_antal = 0
        reset = True
    
    # Oppna alla filer for att kunna lasa och skriva
    

    # Anvands for att undvika felmeddelande nar loop avbryts
    if checking != '0,exit':

        # Leta igenom medlemsregister efter medlem och spara i namn
        for row in range(2, medSheet.max_row + 1):
            if medSheet['A' + str(row)].value == checking:
                namn = medSheet['B' + str(row)].value
                
                # Kolla även om personen är med i styret
                if medSheet['C' + str(row)].value == 'Styret':
                    Styret = True


        # Om kortet inte finns i registret, fråga efter det gamla kortet
        if namn == unknown:
            
            # Starta timer, 5 sekunder
            t4 = Timer(5.0, tid_funk)
            while checking != '0,exit':
                t4.start()
                # Vänta på input från läsaren
                checking3 = input("Now you can scan your old card! Swipe again (or wait 5 seconds) to abort  ")
                t4.cancel()
                t4.join()
                break
            
            checking2 = "0,"+checking3
                
            # Leta i medlemsregistret efter medlem (gamla kortet)
            for row in range(2, medSheet.max_row + 1):
                if medSheet['A' + str(row)].value == checking2:
                    namn = medSheet['B' + str(row)].value
                    row_name = row
                    
                    # Kolla även om personen är med i styret
                    if medSheet['C' + str(row)].value == 'Styret':
                        Styret = True
                    
            # Om medlem hittas, ersätt kortnummret, annars avbryt
            if namn != unknown:
                medSheet['A' + str(row_name)].value = checking
                
        # Om medlemmen hittats kör vidare, annars hoppa vidare
        if namn != unknown:
            
            reset = True
            # Kolla om personen ar inloggad
            for medlem in incheckade:
                if medlem == namn:
                    status = True
                    break
            # Kolla vart info ska loggas samt logga det
            if status:
                for row in range(2, loggSheet.max_row + 1):
                    if loggSheet['B' + str(row)].value == namn and not loggSheet['D' + str(row)].value:
                        loggplats = row                   
            else:
                loggplats = loggSheet.max_row + 1
                loggSheet['B' + str(loggplats)].value = namn # 15374926

            # Kollar så att utcheckning sker samma dag, annars skrivs det upp
            if status and loggSheet['A' + str(loggplats)].value == str(date):
                loggSheet['D' + str(loggplats)].value = str(time_now)
                incheckade.remove(namn)
                if Styret:
                    Styret_antal -= 1
                    
            elif status:
                loggSheet['D' + str(loggplats)].value = str(time_now)
                loggSheet['E' + str(loggplats)].value = str(date)
                loggSheet['F' + str(loggplats)].value = "Late checkout"
                incheckade.remove(namn)
                if Styret:
                    Styret_antal -= 1
                    
            else:
                loggSheet['A' + str(loggplats)].value = str(date)
                loggSheet['C' + str(loggplats)].value = str(time_now)
                incheckade.append(namn)
                if Styret:
                    Styret_antal += 1

        if reset:
            reset = False
            # Anvands for att "kontrollera" incheckningsfonstret
            if alpha:
                #logo.destroy()
                text.destroy()


            item = "\n".join(incheckade)

    ##        logo = Text(root, height=40, width=90)
    ##        #image = Image.open(file='home/pi/XP-loggo.PNG')
    ##        #photo = ImageTK.PhotoImage(image)
    ##        #logo.image_create(END, image=photo)
    ##        logo.pack(side=LEFT)


            text = Text(root, height=50, width=100)
            text.tag_configure('message', font=('Arial', 36, 'bold'))
            text.tag_configure('list', font=('Arial', 14))
            text.tag_configure('rubrik', font=('Arial', 20, 'bold'))
            text.insert(END, '\n')     
            if status:
                text.insert(END, "Goodbye %s\n\n" %namn, 'message')
            else:
                text.insert(END, "Welcome %s\n\n" %namn, 'message')
            text.insert(END, 'Incheckade:\n', 'rubrik')
            text.insert(END, item, 'list')
            text.pack(side=RIGHT)



            root.update_idletasks()

            alpha = True

            if status:
                status = False
                
        Styret = False
        namn = unknown
        medreg.save('Medlemsregister.xlsx')
        loggbok.save('Loggbok.xlsx')
        loggbok.save('info/Loggbok_extern.xlsx')
        
        # Hantera bilder, "antal incheckade"
        if len(incheckade) == 0:
            copyfile('/mnt/www/incheckade/0.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 1:
            copyfile('/mnt/www/incheckade/1.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 2:
            copyfile('/mnt/www/incheckade/2.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 3:
            copyfile('/mnt/www/incheckade/3.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 4:
            copyfile('/mnt/www/incheckade/4.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 5:
            copyfile('/mnt/www/incheckade/5.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 6:
            copyfile('/mnt/www/incheckade/6.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 7:
            copyfile('/mnt/www/incheckade/7.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 8:
            copyfile('/mnt/www/incheckade/8.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 9:
            copyfile('/mnt/www/incheckade/9.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) == 10:
            copyfile('/mnt/www/incheckade/10.png','/mnt/www/incheckade/incheckade.png')
        elif len(incheckade) > 10:
            copyfile('/mnt/www/incheckade/fler.png','/mnt/www/incheckade/incheckade.png')
            
        # Hantera bilder, "antal styret"
        if Styret_antal == 0:
            copyfile('/mnt/www/incheckade/0.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 1:
            copyfile('/mnt/www/incheckade/1.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 2:
            copyfile('/mnt/www/incheckade/2.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 3:
            copyfile('/mnt/www/incheckade/3.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 4:
            copyfile('/mnt/www/incheckade/4.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 5:
            copyfile('/mnt/www/incheckade/5.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 6:
            copyfile('/mnt/www/incheckade/6.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 7:
            copyfile('/mnt/www/incheckade/7.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 8:
            copyfile('/mnt/www/incheckade/8.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 9:
            copyfile('/mnt/www/incheckade/9.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal == 10:
            copyfile('/mnt/www/incheckade/10.png','/mnt/www/incheckade/styret.png')
        elif Styret_antal > 10:
            copyfile('/mnt/www/incheckade/fler.png','/mnt/www/incheckade/styret.png')
    
#Gor variabler for manad och ar, anvand dessa till att skapa nya flikar,
#en for varje manad.
#Vilken tar den som utcheckning om det redan finns en icke/registrerad
#sedan tidigare? Lagg till utcheckningsdatum vid sen utcheckning
#Stall in ratt tidszon
#Om karkort anvands, anvand det nya for att ersatta det gamla i medlemsregistret
#Fixa bild till incheckningen
#Utcheckningstid reggas aven vid sen checkning, lagg till notis vid sidan
