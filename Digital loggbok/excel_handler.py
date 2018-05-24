
from datetime import datetime
import openpyxl
import member

days_saved_online = 30
path_to_logg_online = 'Loggbok_online.xlsx'
path_to_member_register = 'Medlemsregister.xlsx'
path_to_new_members = 'Nyamedlemmar.xlsx'
path_to_datalogger = ''
#path_to_logg_online = 'info/Loggbok_online.xlsx'
#path_to_member_register = '/mnt/www/Medlemsregister.xlsx'
#path_to_new_members = '/mnt/www/Nyamedlemmar.xlsx'
#path_to_datalogger = '/mnt/www/datalog/'
latest_save = datetime.now()
loggbok = openpyxl.load_workbook(path_to_logg_online)
loggSheet = loggbok.active



def init_member_register():
    #global path_to_member_register
    medreg = openpyxl.load_workbook(path_to_member_register)
    medSheet = medreg[medreg.sheetnames[0]]
    for row in range(2, medSheet.max_row + 1):
        keyCard = medSheet['A' + str(row)].value
        name =  medSheet['B' + str(row)].value
        board_member = medSheet['C' + str(row)].value == 'Styret'
        member.Member(keyCard, name, board_member)




def clean_earliest_loggbook(self):
    date_today = datetime.now()
    removed_rows = 0
    for row in range(2, loggSheet.max_row):
        time = loggSheet['A'+str(row)].value
        if time == None or \
        (date_today - strptime(time,"%Y-%m-%d")).days > days_saved_online:
            loggSheet.delete_rows(row-idx_removed, 1)
            idx_removed = idx_removed + 1

def save():
        loggbok.save(path_to_logg_online)
        # ERROR ERROR ERROR FIX THIS
        current_month = datetime.now().strftime('%Y%B')
        loggbok.save('%s%s.xlsx' %(path_to_datalogger, current_month))
        global latest_save
        latest_save = datetime.now()

def cardreader_parser(cardkey_str):
    cardreader_bits = 24
    cardkey_binary = bin(int(cardkey_str))
    cardkey_binary_appended = '0'*cardreader_bits + cardkey_binary
    return ('0,%i' %int(cardkey_binary_appended[-cardreader_bits:], 2))

def import_new_members():
    # Använd bara denna på natten eller något.
    medreg = openpyxl.load_workbook(path_to_new_members)
    medSheet = medreg[medreg.sheetnames[0]]
    for row in range(2, medSheet.max_row + 1):
        keyCard = cardreader_parser(medSheet['A' + str(row)].value)
        name =  medSheet['B' + str(row)].value
        board_member = medSheet['C' + str(row)].value == 'Styret'
        Member(keyCard, name, board_member)
    medreg.close()
    medreg = openpyxl.Workbook()
    medSheet = medreg.active
    medSheet.title = 'Nya medlemmar'
    medSheet['A1'] = 'Nyckelnr'
    medSheet['B1'] = 'Namn'
    medSheet['C1'] = 'Styrelsemedlem'
    medreg.save(path_to_new_members)
    medreg.close()
    save_memberlist_to_file()

def save_memberlist_to_file():
    medreg = openpyxl.Workbook()
    medSheet = medreg.active
    medSheet.title = 'Medlemsregister'
    medSheet['A1'] = '0,Nyckelnr'
    medSheet['B1'] = 'Namn'
    medSheet['C1'] = 'Styrelsemedlem'
    row = 2
    for member in Member.member_register:
        medSheet['A' + str(row)] = member.key_card
        medSheet['B' + str(row)] = member.name
        if member.board_member:
            medSheet['C' + str(row)] = 'Styret'
    medreg.save(path_to_member_register)
    medreg.close()

def save_all_checkedin_to_log():
    for members in Member.checked_in_members:
        row = str(loggSheet.max_row + 1)
        loggSheet['A' + row] = members.getCheckinDate()
        loggSheet['B' + row] = members.getName()
        loggSheet['C' + row] = members.getCheckedInTime()
    for members in Member.checked_in_styret:
        row = str(loggSheet.max_row + 1)
        loggSheet['A' + row] = members.getCheckinDate()
        loggSheet['B' + row] = members.getName()
        loggSheet['C' + row] = members.getCheckedInTime()
    latest_save = datetime.now()
    save()

def save_to_log(member):
    time_now_str = datetime.now().strftime("%H:%M:%S")
    row = str(loggSheet.max_row + 1)
    loggSheet['A' + row] = member.getCheckinDate()
    loggSheet['B' + row] = member.getName()
    loggSheet['C' + row] = member.getCheckedInTime()
    loggSheet['D' + row] = time_now_str
    if time_now_str < member.getCheckedInTime():
        loggSheet['E' + row] = "Late checkout"







     
